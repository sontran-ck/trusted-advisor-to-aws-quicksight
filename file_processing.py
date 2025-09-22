import pandas as pd
import re
import json
import copy

import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
class ExcelProcessing:
    def __init__(self):
        pass


    def read_file(self):
        with open(self.file_path, 'r') as file:
            return file.read()

    def write_file(self, data):
        with open(self.file_path, 'w') as file:
            file.write(data)

    def handle_rows(self,sheet_name,df_sheet, null_value=None):
        
        title = df_sheet.iat[0,0]
        account_id = null_value
        real_description = null_value
        status = null_value
        total_number_of_resources_processed = null_value
        number_of_resources_flagged = null_value
        number_of_suppressed_resources = null_value
        source = null_value
        alert_criteria = null_value
        recommended_action = null_value
        additional_resources = null_value
        
        if "AWS Account ID:" in df_sheet.iat[1,0]:
            account_id = df_sheet.iat[1,0].partition(':')[2].strip()
        if "Description:" in df_sheet.iat[2,0]:
            description_text = df_sheet.iat[2, 0]

            # Tách phần mô tả chính
            real_description_match = re.split(r"\n\n(Source|Alert Criteria|Recommended Action|Additional Resources)", description_text, maxsplit=1)
            real_description = real_description_match[0].partition(":")[2].strip()

            # Chuẩn bị biến kết quả
            source = recommended_action = additional_resources = null_value
            alert_criteria = null_value

            # Regex pattern tìm từng section
            pattern = re.compile(
                r"(Source|Alert Criteria|Recommended Action|Additional Resources)\s*(.*?)\s*(?=(Source|Alert Criteria|Recommended Action|Additional Resources|$))",
                re.DOTALL,
            )

            for match in pattern.finditer(description_text):
                section, content = match.group(1), match.group(2).strip()
                if section == "Source":
                    source = content.strip()
                elif section == "Alert Criteria":
                    alert_criteria = content.strip()
                elif section == "Recommended Action":
                    recommended_action = content.strip()
                elif section == "Additional Resources":
                    additional_resources = content.strip()

        if "Status:" in df_sheet.iat[3,0]:
            status = df_sheet.iat[3,0].partition(':')[2].strip()
        if "Total number of resources processed:" in df_sheet.iat[5,1]:
            total_number_of_resources_processed = df_sheet.iat[5,1].partition(':')[2].strip()
        if "Number of resources flagged:" in df_sheet.iat[6,1]:
            number_of_resources_flagged = df_sheet.iat[6,1].partition(':')[2].strip()
        if "Number of suppressed resources:" in df_sheet.iat[7,1]:
            number_of_suppressed_resources = df_sheet.iat[7,1].partition(':')[2].strip()

        if status == "not_available":
            return {
                "check_title": title,
                # "account_id": account_id,
                "description": null_value,
                "total_number_of_resources_processed": null_value,
                "number_of_resources_flagged": null_value,
                "number_of_suppressed_resources": null_value,
                "source": null_value,
                "alert_criteria": null_value,
                "recommended_action": null_value,
                "additional_resources": null_value,
                "status": "not_available",
            }
        return {
            "check_title": title,
            # "account_id": account_id,
            "description": real_description,
            "total_number_of_resources_processed": total_number_of_resources_processed,
            "number_of_resources_flagged": number_of_resources_flagged,
            "number_of_suppressed_resources": number_of_suppressed_resources,
            "source": source,
            "alert_criteria": alert_criteria,
            "recommended_action": recommended_action,
            "additional_resources": additional_resources,
            "status": status
        }

    def excel_to_json(self, file_path):
        self.file_path = file_path
        try: 
            df = pd.read_excel(file_path, sheet_name=None, header=None)
        except Exception as e:
            print(f"Error reading excel file: {e}")
            raise e
        output = []
        for sheet_name, df_sheet in df.items():
            output.append(self.handle_rows(sheet_name,df_sheet, None))
        return output

    def flatten_json_data(self, data):
        """
        Flatten JSON data - tách alert_criteria thành các records riêng biệt
        
        Args:
            data: List of dictionaries từ excel_to_json()
        
        Returns:
            List of flattened dictionaries
        """
        flattened_data = []
        
        for record in data:
            if record.get('alert_criteria'):
                # Tách alert_criteria thành các record riêng biệt
                base_record = record.copy()
                alert_criteria_list = base_record.pop('alert_criteria', [])
                
                if alert_criteria_list:
                    # Tạo record cho mỗi alert criterion
                    for criterion in alert_criteria_list:
                        flattened_record = base_record.copy()
                        flattened_record['alert_level'] = criterion.get('level')
                        flattened_record['alert_description'] = criterion.get('description')
                        flattened_data.append(flattened_record)
                else:
                    # Nếu không có alert_criteria thì giữ nguyên record
                    base_record['alert_level'] = None
                    base_record['alert_description'] = None
                    flattened_data.append(base_record)
            else:
                # Nếu không có alert_criteria thì thêm fields null và giữ nguyên
                record_copy = record.copy()
                record_copy['alert_level'] = None
                record_copy['alert_description'] = None
                flattened_data.append(record_copy)
        
        return flattened_data

    def convert_to_ndjson(self, data, output_file=None, flatten_alert_criteria=False):
        """
        Convert data to NDJSON format (Newline Delimited JSON)
        
        Args:
            data: List of dictionaries từ excel_to_json()
            output_file: Tên file để save, nếu None thì return string
            flatten_alert_criteria: Nếu True, tách alert_criteria thành các dòng riêng biệt
        
        Returns:
            String NDJSON hoặc save to file
        """
        if flatten_alert_criteria:
            # Sử dụng function flatten_json_data
            data_to_process = self.flatten_json_data(data)
        else:
            # Giữ nguyên format gốc
            data_to_process = data
        
        ndjson_lines = []
        for record in data_to_process:
            ndjson_lines.append(json.dumps(record, ensure_ascii=False))
        
        ndjson_content = '\n'.join(ndjson_lines)
        
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(ndjson_content)
            print(f"NDJSON data saved to {output_file}")
            return output_file
        else:
            return ndjson_content

    def save_json(self, data, file_path):
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)



class JsonProcessing:
    def __init__(self):
        pass

    def read_json_file(self, file_path):
        self.file_path = file_path
        with open(self.file_path, 'r', encoding='utf-8') as f:
            self.pd_data = pd.read_json(f)

    
    def save_excel(self, data, output_file):
        pd_data = pd.DataFrame(data)
        pd_data.to_excel(output_file, sheet_name='Table 1', index=False)

    def parse_alert_criteria(self, alert_criteria_text):
        """
        Parse alert criteria text into structured JSON format
        Input: "Red: description\nYellow: description\nGreen: description\nBlue: description"
        Output: [{"level": "Red", "description": "description"}, ...]
        """
        if not alert_criteria_text or alert_criteria_text.strip() == "":
            return []
        
        criteria_list = []
        lines = alert_criteria_text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Tìm pattern "Color:" ở đầu dòng
            color_match = re.match(r'^(Red|Yellow|Green|Blue):\s*(.+)$', line, re.IGNORECASE)
            if color_match:
                level = color_match.group(1).capitalize()
                description = color_match.group(2).strip()
                criteria_list.append({
                    "level": level,
                    "description": description
                })
        
        return criteria_list

    def get_parse_data(self, data):
        copy_data = copy.deepcopy(data)
        for record in copy_data:
            if record.get('alert_criteria'):
                record['alert_criteria'] = self.parse_alert_criteria(record['alert_criteria'])
        return copy_data

    def summary_alert_criteria(self, parse_data):
        """
        Summary alert criteria
        Args:
            parse_data: List of dictionaries of parse data 
        Returns:
            Summary alert criteria
        """
        summary_alert_criteria = {
            "Green": 0,
            "Blue": 0,
            "Yellow": 0,
            "Red": 0,
            # "Total": 0
        }
        for record in parse_data:
            if record.get('alert_criteria'):
                for criterion in record['alert_criteria']:
                    summary_alert_criteria[criterion['level']] += 1
                    # summary_alert_criteria['Total'] += 1
        return summary_alert_criteria

    def insert_data_detail_alert_criteria(self, extracted_json_data, null_value = None):
        """
        Insert data detail alert criteria
        Args:
            extracted_json_data: List of dictionaries of extracted json data
        Returns:
            Insert data detail alert criteria
        """
        output = []
        parse_data = self.get_parse_data(extracted_json_data)
        
        copy_data = copy.deepcopy(extracted_json_data)
        for index, record in enumerate(copy_data):
            record['green_alert_criteria'] = null_value
            record['blue_alert_criteria'] = null_value
            record['yellow_alert_criteria'] = null_value
            record['red_alert_criteria'] = null_value
            if parse_data[index].get('alert_criteria'):
                record['green_alert_criteria'] = False
                record['blue_alert_criteria'] = False
                record['yellow_alert_criteria'] = False
                record['red_alert_criteria'] = False
                for criterion in parse_data[index]['alert_criteria']:
                    level = criterion['level']
                    if level == 'Green':
                        record['green_alert_criteria'] = True
                    elif level == 'Yellow':
                        record['yellow_alert_criteria'] = True
                    elif level == 'Red':
                        record['red_alert_criteria'] = True
                    elif level == 'Blue':
                        record['blue_alert_criteria'] = True

            output.append(record)
        return output

    def create_chart_alert_criteria(summary_alert_criteria, file_path, output_file):
        workbook = load_workbook(file_path)
        # append sheet
        sheet = workbook.create_sheet("Summary Alert Criteria")
        sheet.cell(row=1, column=1).value = "Alert Criteria"
        sheet.cell(row=1, column=2).value = "Count"
        for i, (key, value) in enumerate(summary_alert_criteria.items(), start=2):
            sheet.cell(row=i, column=1).value = key
            sheet.cell(row=i, column=2).value = value
        # Tạo pie chart với matplotlib để có thể thiết lập màu sắc chính xác
        labels = list(summary_alert_criteria.keys())
        values = list(summary_alert_criteria.values())
        
        # Thiết lập màu sắc phù hợp với tên key
        color_mapping = {
            "Red": "#FF0000",      # Đỏ
            "Blue": "#0000FF",     # Xanh dương  
            "Green": "#00FF00",    # Xanh lá
            "Yellow": "#FFFF00"    # Vàng
        }
        
        # Tạo danh sách màu sắc theo thứ tự của labels
        colors = []
        for label in labels:
            if label in color_mapping:
                colors.append(color_mapping[label])
            else:
                colors.append("#808080")  # Màu xám mặc định cho key không có trong mapping
        
        # Tạo pie chart với matplotlib
        plt.figure(figsize=(8, 6))
        plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        plt.title("Alert Criteria")
        plt.axis('equal')  # Đảm bảo biểu đồ tròn
        
        # Lưu biểu đồ thành file ảnh
        chart_filename = "alert_criteria_pie_chart.png"
        plt.savefig(chart_filename, dpi=300, bbox_inches='tight')
        plt.close()
        
        # Chèn ảnh vào Excel
        img = Image(chart_filename)
        img.width = 400  # Điều chỉnh kích thước
        img.height = 300
        sheet.add_image(img, "D1")
        
        print(f"Pie chart created with colors: {dict(zip(labels, colors))}")
        
        workbook.save(output_file)
        
        # Dọn dẹp file ảnh tạm thời
        
        if os.path.exists(chart_filename):
            os.remove(chart_filename)
            print(f"Temporary chart file {chart_filename} removed")